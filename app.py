# -*- coding: utf-8 -*-
"""
독서활동기록 ISBN 검증 대시보드
--------------------------------
교사가 학생들의 독서기록 엑셀(.xlsx)을 업로드하면,
국립중앙도서관 서지정보 API와 대조하여 공식 등재 도서만 정제해
새로운 엑셀 파일로 반환하는 Streamlit 웹 애플리케이션입니다.

실행:  streamlit run app.py
"""

import io
import os
import re
import time
import urllib.parse

import pandas as pd
import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 상수 / 설정
# ---------------------------------------------------------------------------
API_URL = "https://www.nl.go.kr/seoji/SearchApi.do"

ERROR_MESSAGE = "오류: ISBN 미등재 또는 도서 정보 불일치"
REQUEST_TIMEOUT = 30          # API 요청 타임아웃(초) - 서버가 느릴 때 대비
REQUEST_RETRIES = 3           # 타임아웃 시 자동 재시도 횟수
REQUEST_DELAY = 0.3           # 호출 간 대기(초) - 과도한 요청 방지
PAGE_SIZE = 50               # API 검색 결과 페이지 크기(누락 방지 위해 넉넉히)

st.set_page_config(page_title="독서기록 ISBN 검증 대시보드", page_icon="📚", layout="wide")


# ---------------------------------------------------------------------------
# 유틸리티 함수
# ---------------------------------------------------------------------------
def strip_html(text) -> str:
    """응답에 섞인 <span ...>...</span> 같은 HTML 태그와 엔티티를 제거."""
    if text is None:
        return ""
    s = str(text)
    s = re.sub(r"<[^>]+>", "", s)           # 태그 제거
    s = s.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    s = s.replace("&quot;", '"').replace("&#39;", "'").replace("&nbsp;", " ")
    return s.strip()


def display_title(title: str) -> str:
    """화면 표시용 제목: 앞 괄호는 그대로 두고, 콜론(:) 부제와 '뒤쪽' 괄호만 제거.
    예) '(영원한 라이벌) 김대중 VS 김영삼' → 그대로 (앞 괄호 유지)
        '검은 사슴 (한국문학전집 024)'      → '검은 사슴' (뒤 괄호 제거)
        '페인트 : 이희영 장편소설'          → '페인트'
    """
    s = strip_html(title)
    s = re.split(r"\s*[:：]\s*", s)[0]
    # 끝에 붙은 괄호 묶음 반복 제거 (시리즈/판본 표기 등)
    prev = None
    while prev != s:
        prev = s
        s = re.sub(r"\s*[\(\[（［〔《〈][^)\]）］〕》〉]*[\)\]）］〕》〉]\s*$", "", s)
    return re.sub(r"\s+", " ", s).strip()


def strip_subtitle(title: str) -> str:
    """제목에서 괄호 묶음과 부제(콜론 ' : ' 뒤)를 제거해 원제만 남긴다.
    예) '페인트 : 이희영 장편소설'        → '페인트'
        '(영원한 라이벌) 김대중 VS 김영삼' → '김대중 VS 김영삼'
        '어린왕자 (The Little Prince)'     → '어린왕자'
    """
    s = strip_html(title)
    # 괄호 묶음 제거: ( ), [ ], （ ）, ［ ］, 〔 〕, 《 》, < >
    s = re.sub(r"[\(\[（［〔《〈<][^)\]）］〕》〉>]*[\)\]）］〕》〉>]", " ", s)
    # 콜론 부제 제거
    s = re.split(r"\s*[:：]\s*", s)[0]
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _title_key(t: str) -> str:
    """제목 비교용 키: 부제·괄호 제거 + 끝 권수 제거 + 공백 정규화."""
    s = strip_subtitle(t).lower()
    s = re.sub(r"[^\w가-힣 ]", "", s)   # 한글/영숫자/공백만
    # 끝에 붙은 권수 표기 제거: ' 1', ' 2권', ' 3편', ' 상/중/하' 등
    s = re.sub(r"\s+(\d{1,4}\s*(권|편|화|부|집|쇄|화권)?|[상중하])$", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def title_relation(in_title: str, doc_title: str):
    """입력 제목과 결과 제목의 관계를 판정.
    반환: 'exact' | None
      - exact : 부제·괄호·끝 권수 제거 후 (띄어쓰기 무시) 완전히 같음
      - None  : '시한부' vs '시한부 연애', '국가' vs '국가론' 등 다른 제목은 불인정
    (권수 차이 '토지'/'토지 1'는 제목 키 단계에서 권수를 떼므로 exact로 처리됨)
    """
    a, b = _title_key(in_title), _title_key(doc_title)
    if not a or not b:
        return None
    if a.replace(" ", "") == b.replace(" ", ""):   # 띄어쓰기 차이는 같은 제목으로
        return "exact"
    return None


def normalize(text) -> str:
    """공백/특수문자를 제거하고 소문자화하여 비교용 문자열로 정규화."""
    if text is None:
        return ""
    s = strip_html(text)                       # HTML 태그 먼저 제거
    s = re.sub(r"[\s\u200b]+", "", s)          # 공백 및 zero-width space 제거
    s = re.sub(r"[^\w가-힣]", "", s)           # 한글/영숫자 외 기호 제거
    return s.lower()


def get_stored_key() -> str:
    """
    숨겨둔 인증키를 안전하게 읽어온다.
    우선순위: st.secrets["NL_API_KEY"]  ->  환경변수 NL_API_KEY
    (둘 다 없으면 빈 문자열을 반환하고, 화면에서 직접 입력받는다.)
    """
    # st.secrets 는 secrets 파일이 없으면 예외를 던지므로 방어적으로 접근
    try:
        if "NL_API_KEY" in st.secrets:
            return str(st.secrets["NL_API_KEY"]).strip()
    except Exception:  # noqa: BLE001
        pass
    return os.environ.get("NL_API_KEY", "").strip()


# 저자(author) 역할어와 비저자(translator/editor 등) 역할어
_AUTHOR_ROLES = ["지은이", "지음", "저자", "원작", "공저", "글", "著", "저"]
_OTHER_ROLES = ["옮긴이", "옮김", "역자", "번역", "영역", "한역", "譯", "역",
                "펴낸이", "엮은이", "엮음", "일러스트", "그린이", "그림", "삽화",
                "사진", "편집", "곁텍스트", "감수", "구성", "편저", "편", "감독",
                "기획", "해설", "캘리그래피", "낭독자", "낭독", "녹음", "연주",
                "노래", "출연", "진행", "원작자", "각색", "번안", "주해", "주석",
                "교정", "교열", "디자인", "표지", "장정"]
_ALL_ROLES = sorted(_AUTHOR_ROLES + _OTHER_ROLES, key=len, reverse=True)
_ROLE_SPLIT_RE = re.compile(r"(.+?)\s+(" + "|".join(map(re.escape, _ALL_ROLES)) + r")(?=[\s;,/·]|$)")
_ORG_RE = re.compile(r"(출판사|출판|편집부|편찬|연구소|연구회|위원회|미디어|에디터스|"
                     r"주식회사|\(주\)|문화원|박물관|진흥원|아카데미|컴퍼니|컴퍼니|스튜디오)")
_HAS_COLON_ROLE_RE = re.compile(r"[^\s:：]+\s*[:：]")


def _is_author_role(role: str) -> bool:
    """역할어 문자열이 '저자'에 해당하는지 판정(번역/삽화 등은 제외)."""
    if any(o in role for o in _OTHER_ROLES):
        return False
    return any(a in role for a in _AUTHOR_ROLES)


def _strip_edge_roles(nm: str) -> str:
    """이름 끝에 토큰으로 붙은 역할어(글, 그림, 글·그림 등)와 기호를 제거."""
    changed = True
    while changed:
        changed = False
        nm = nm.strip(" .,;/·")
        for role in _ALL_ROLES:
            if nm.endswith(role) and len(nm) > len(role):
                prev = nm[: -len(role)]
                if prev[-1] in " .,;/·":      # 토큰 경계일 때만 제거(이름 글자 보호)
                    nm = prev.strip(" .,;/·")
                    changed = True
                    break
    return nm


def _clean_name(nm: str) -> str:
    """이름 하나를 정리: 생몰년[..] 제거, 끝 역할어 제거, '성, 이름' → '이름 성'."""
    nm = re.sub(r"\[[^\]]*\]", "", nm)          # [1877-1962] 등 제거
    nm = re.sub(r"\s+", " ", nm).strip(" .,;/·")
    nm = _strip_edge_roles(nm)                   # 끝에 남은 '글·그림' 등 제거
    if "," in nm:                                # 'Hesse, Hermann' / '헤세, 헤르만'
        parts = [p.strip() for p in nm.split(",") if p.strip()]
        if len(parts) == 2:
            nm = f"{parts[1]} {parts[0]}".strip()
    return nm.strip()


def clean_author(author_raw: str) -> str:
    """
    저자 필드를 사람이 읽기 좋은 형태로 정리. 두 가지 표기를 모두 처리.
      (A) '이름 역할 이름 역할 …'      예) Hesse, Hermann 지은이 …
      (B) '역할: 이름 ; 역할: 이름 …'  예) 지은이: 생텍쥐페리 ;영역: 제니 박 …
    저자 역할(지은이/지음/글 등)만 남기고, 여러 명이면 '첫저자 외 N인'.
    """
    if not author_raw:
        return ""
    s = strip_html(author_raw)

    authors = []   # 지은이/지음/글 등
    others = []    # 엮음/옮김/펴낸이 등 (저자가 없을 때 대비)
    matched = False

    if _HAS_COLON_ROLE_RE.search(s):
        # (B) 역할: 이름 형식 — ';' 또는 '/'로 그룹 구분
        for group in re.split(r"[;/]", s):
            group = group.strip()
            if not group:
                continue
            mm = re.match(r"\s*([^:：]+)[:：]\s*(.+)$", group)
            if not mm:
                continue
            matched = True
            role, names = mm.group(1).strip(), mm.group(2).strip()
            bucket = authors if _is_author_role(role) else others
            for nm in re.split(r"\s*[,·]\s*", names):   # 공동저자 분리
                nm = _clean_name(nm)
                if nm:
                    bucket.append(nm)
    else:
        # (A) 이름 역할 형식
        for m in _ROLE_SPLIT_RE.finditer(s):
            matched = True
            name, role = m.group(1).strip(), m.group(2)
            bucket = authors if _is_author_role(role) else others
            nm = _clean_name(name)
            if nm:
                bucket.append(nm)

    if not matched:                            # 역할어가 전혀 없음
        # '이름[생몰년] 이름[생몰년] …' 형태면 여러 명으로 보고 정리
        bracket_names = [n.strip() for n in re.findall(r"([^\[\]\s][^\[\]]*?)\s*\[", s)]
        bracket_names = [n for n in bracket_names if n]
        if len(bracket_names) >= 2:
            first = _clean_name(bracket_names[0])
            return f"{first} 외 {len(bracket_names) - 1}인" if first else ""
        # '강문일,김경진,…' 처럼 콤마/가운뎃점/세미콜론으로 여러 명 나열된 경우
        parts = [_clean_name(p) for p in re.split(r"[,·;]", s)]
        parts = [p for p in parts if len(p) >= 2]
        non_org = [p for p in parts if not _ORG_RE.search(p)]
        if non_org:                            # 개인 저자가 있으면 기관명은 제외
            parts = non_org
        if len(parts) >= 2:
            return f"{parts[0]} 외 {len(parts) - 1}인"
        if len(parts) == 1:
            return parts[0]
        return _clean_name(s)

    # 지은이가 없으면 엮은이/옮긴이 등 기타 기여자를 대신 사용(선집·번역서 등)
    pool = authors if authors else others

    # 중복 제거(순서 유지)
    seen, uniq = set(), []
    for a in pool:
        if a not in seen:
            seen.add(a)
            uniq.append(a)
    pool = uniq

    if not pool:
        return ""
    if len(pool) == 1:
        return pool[0]
    return f"{pool[0]} 외 {len(pool) - 1}인"


# ---------------------------------------------------------------------------
# 국립중앙도서관 API 연동
# ---------------------------------------------------------------------------
ERROR_CODE_MEANING = {
    "000": "시스템 오류",
    "010": "인증키 누락",
    "011": "유효하지 않은 인증키(승인 여부 확인 필요)",
    "012": "필수 파라미터 누락",
}


def extract_records(data) -> list:
    """JSON 응답에서 도서 레코드 리스트를 견고하게 찾아낸다."""
    if isinstance(data, list):
        return data
    if not isinstance(data, dict):
        return []
    for key in ("result", "docs", "items", "list", "RESULT", "resultList"):
        v = data.get(key)
        if isinstance(v, list):
            return v
    # 못 찾으면: 값들 중 dict들의 리스트를 결과로 간주
    for v in data.values():
        if isinstance(v, list) and v and isinstance(v[0], dict):
            return v
    return []


@st.cache_data(show_spinner=False)
def call_seoji_api(query: str, cert_key: str, is_isbn: bool = False) -> dict:
    """
    국립중앙도서관 서지정보(Seoji) DB 검색. is_isbn=True면 ISBN으로, 아니면 제목으로.
    반환 dict: docs / error / raw
    """
    q = str(query).strip()
    if not q:
        return {"docs": [], "error": None, "raw": ""}

    # 검색어를 직접 UTF-8 퍼센트 인코딩하여 URL을 구성한다.
    # (requests 자동 인코딩이 일부 한글 음절에서 서버와 어긋나 011을 유발하는 문제 회피)
    parts = [
        f"cert_key={urllib.parse.quote(str(cert_key), safe='')}",
        "result_style=json",
        "page_no=1",
        f"page_size={PAGE_SIZE}",
    ]
    if is_isbn:
        parts.append(f"isbn={urllib.parse.quote(re.sub(r'[^0-9Xx]', '', q), safe='')}")
    else:
        parts.append(f"title={urllib.parse.quote(q, safe='')}")
    request_url = API_URL + "?" + "&".join(parts)
    headers = {"User-Agent": "Mozilla/5.0 (reading-checker)"}

    last_err = None
    raw = ""
    data = None
    for attempt in range(REQUEST_RETRIES):
        try:
            resp = requests.get(request_url, headers=headers, timeout=REQUEST_TIMEOUT)
            resp.raise_for_status()
            raw = resp.text[:800]
            data = resp.json()
            last_err = None
            break
        except requests.exceptions.Timeout:
            last_err = "서버 응답이 느립니다(시간 초과). 잠시 후 다시 시도됩니다."
            time.sleep(1.0)  # 잠깐 쉬었다가 재시도
            continue
        except requests.RequestException as e:
            return {"docs": [], "error": f"네트워크 오류: {e}", "raw": ""}
        except ValueError:
            # JSON이 아님 → 보통 인증 오류 안내 페이지가 돌아온 경우
            return {"docs": [], "error": "응답을 해석할 수 없습니다(인증키/주소 확인).", "raw": raw}

    if data is None:
        return {"docs": [], "error": last_err or "알 수 없는 오류", "raw": ""}

    # 공식 에러코드 처리
    code = str(
        data.get("ERR_CODE") or data.get("error_code")
        or data.get("ERROR_CODE") or data.get("CODE") or ""
    ).strip() if isinstance(data, dict) else ""
    if code and code in ERROR_CODE_MEANING:
        return {"docs": [], "error": f"[{code}] {ERROR_CODE_MEANING[code]}", "raw": raw}

    docs = extract_records(data)
    return {"docs": docs, "error": None, "raw": raw}


def match_book(title: str, author: str, docs: list):
    """
    검색 결과(docs) 중 입력 제목/저자와 가장 잘 맞는 항목을 선택.
    반환: (matched_doc | None, score)
      - 제목 일치(부분 포함)는 필수
      - 저자 일치 시 가중치 부여
    """
    n_author = normalize(author)
    if not _title_key(title):
        return None, 0

    best_doc, best_score = None, 0
    for doc in docs:
        rel = title_relation(title, doc.get("TITLE", ""))
        if rel is None:                 # 중간 부분일치(페인트회사 등)는 제외
            continue
        # 종이책만 인정 (전자책·오디오북·판형 미표기는 제외)
        if str(doc.get("FORM", "")).strip() != "종이책":
            continue
        doc_author = normalize(clean_author(doc.get("AUTHOR", "")))

        # 제목 점수: 정확일치를 가장 강하게(ISBN보다 우선)
        score = 5 if rel == "exact" else 2

        # 저자 일치 판정 (입력 저자가 있을 때만 평가)
        if n_author:
            if doc_author and (n_author in doc_author or doc_author in n_author):
                score += 3
            else:
                score -= 2  # 저자 불일치 감점

        # 저자 정보가 아예 없는 판본은 후순위로(‘저자 미상’ 방지)
        if not doc_author:
            score -= 3
        # 저자 자리가 출판사·기관명인 판본도 후순위로(개인 저자 우선)
        elif _ORG_RE.search(clean_author(doc.get("AUTHOR", ""))):
            score -= 2

        # ISBN 보유 가산점(정확일치보다 약하게)
        if str(doc.get("EA_ISBN", "")).strip():
            score += 2

        if score > best_score:
            best_doc, best_score = doc, score

    return best_doc, best_score


def _matching_candidates(title, docs):
    """제목이 일치하는 레코드들의 (저자) 후보 목록을 만든다.
    같은 책의 중복 레코드는 ISBN/저자명으로 합치고, 저자 미상은 제외한다."""
    cands = []
    seen_book = set()      # ISBN 기준 중복 책 제거
    seen_author = set()    # 같은 저자명 중복 제거
    for d in docs:
        if title_relation(title, d.get("TITLE", "")) is None:
            continue
        if str(d.get("FORM", "")).strip() != "종이책":   # 종이책만
            continue
        book = str(d.get("EA_ISBN", "")).strip()
        if book and book in seen_book:
            continue
        a = clean_author(d.get("AUTHOR", ""))
        if not a or a in seen_author:
            continue
        if book:
            seen_book.add(book)
        seen_author.add(a)
        cands.append(a)
    return cands


def verify_row(title: str, author: str, cert_key: str) -> dict:
    """한 행(도서)을 검증하여 결과 dict 반환."""
    res = call_seoji_api(title, cert_key)

    # 호출/인증 자체가 실패한 경우 → '미등재'와 구분해서 표시
    if res["error"]:
        return {
            "검증결과": "오류",
            "정제된 도서정보": f"API 호출 오류: {res['error']}",
            "ISBN": "",
            "출판사": "",
            "다른 후보": "",
        }

    docs = res["docs"]
    eff_title = title
    doc, score = match_book(title, author, docs)
    success = doc is not None and score >= 1

    # 콤마가 제목의 일부였을 가능성: '제목, 저자'로 잘못 쪼개졌다면
    # 둘을 합쳐 통째 제목으로 한 번 더 검색해 본다.
    if not success and str(author).strip():
        combined = f"{title} {author}".strip()
        res2 = call_seoji_api(combined, cert_key)
        if not res2["error"]:
            doc2, score2 = match_book(combined, "", res2["docs"])
            if doc2 is not None and score2 >= 1:
                doc, score, docs, eff_title = doc2, score2, res2["docs"], combined
                success = True

    if success:
        clean_title = display_title(doc.get("TITLE", eff_title)) or str(eff_title).strip()
        # 표시 저자: 서지정보 DB의 정제된 저자를 기본으로 사용.
        # (입력 저자는 '검색 정확도'를 높이는 데만 쓰고, 표시에는 그대로 쓰지 않음)
        api_auth = clean_author(doc.get("AUTHOR", ""))
        clean_auth = api_auth or "저자 미상"
        isbn = str(doc.get("EA_ISBN", "")).strip()
        publisher = strip_html(doc.get("PUBLISHER", "")).strip().rstrip(" :,")

        # 동명이서 후보: 선택된 저자와 다른 후보들만
        others = [c for c in _matching_candidates(eff_title, docs) if c != clean_auth]
        others = others[:5]

        return {
            "검증결과": "성공",
            "정제된 도서정보": f"{clean_title}({clean_auth})",
            "ISBN": isbn,
            "출판사": publisher,
            "다른 후보": " / ".join(others),
        }
    return {
        "검증결과": "실패",
        "정제된 도서정보": ERROR_MESSAGE,
        "ISBN": "",
        "출판사": "",
        "다른 후보": "",
    }


def verify_isbn(isbn: str, cert_key: str) -> dict:
    """ISBN 하나를 검증하여 결과 dict 반환."""
    res = call_seoji_api(isbn, cert_key, is_isbn=True)
    if res["error"]:
        return {"검증결과": "오류", "정제된 도서정보": f"API 호출 오류: {res['error']}",
                "ISBN": "", "출판사": ""}
    docs = res["docs"]
    if not docs:
        return {"검증결과": "실패", "정제된 도서정보": "오류: 해당 ISBN을 찾을 수 없음",
                "ISBN": "", "출판사": ""}
    doc = docs[0]   # ISBN 검색은 보통 정확히 1건
    clean_title = display_title(doc.get("TITLE", "")) or "(제목 미상)"
    clean_auth = clean_author(doc.get("AUTHOR", "")) or "저자 미상"
    found_isbn = str(doc.get("EA_ISBN", "")).strip() or re.sub(r"[^0-9Xx]", "", str(isbn))
    publisher = strip_html(doc.get("PUBLISHER", "")).strip().rstrip(" :,")
    return {
        "검증결과": "성공",
        "정제된 도서정보": f"{clean_title}({clean_auth})",
        "ISBN": found_isbn,
        "출판사": publisher,
    }
def build_excel(df: pd.DataFrame) -> bytes:
    """결과 DataFrame을 스타일이 적용된 엑셀 바이트로 변환."""
    wb = Workbook()
    ws = wb.active
    ws.title = "검증결과"

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    success_fill = PatternFill("solid", fgColor="E2EFDA")
    fail_fill = PatternFill("solid", fgColor="FCE4E4")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = list(df.columns)

    # 헤더
    for c_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # 데이터
    result_col = columns.index("검증결과") if "검증결과" in columns else None
    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        is_fail = result_col is not None and str(row.iloc[result_col]) == "실패"
        row_fill = fail_fill if is_fail else success_fill
        for c_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
            cell.fill = row_fill
            cell.alignment = left
            cell.border = border

    # 열 너비 자동 조정
    for c_idx, col in enumerate(columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(12, max_len + 4), 55)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# 직접 입력 파싱 + 공통 검증 처리
# ---------------------------------------------------------------------------
def _looks_like_author(s: str) -> bool:
    """콤마 뒤 문자열이 '저자'처럼 보이는지(문장/부제가 아닌지) 판정."""
    s = s.strip()
    if not s or len(s) > 20:
        return False
    if any(ch in s for ch in "?!？！…~"):   # 물음표·느낌표 등이 있으면 제목 일부로 봄
        return False
    if len(s.split()) > 3:                    # 단어가 너무 많으면 저자가 아님
        return False
    return True


def parse_pasted(text: str) -> pd.DataFrame:
    """
    여러 줄 텍스트를 (책 제목, 저자) 표로 변환.
    구분 규칙(제목에 콤마가 들어간 책을 보호):
      1) 탭이 있으면 탭으로 제목/저자 구분 (엑셀에서 두 열 복사·붙여넣기)
      2) 줄 끝이 '(저자)'로 끝나면 괄호 안을 저자로 분리
         - 예) '커튼콜(조우리)' → 제목=커튼콜, 저자=조우리
      3) 탭이 없고 콤마가 있으면, 콤마 뒤가 '저자처럼 보일 때만' 저자로 분리
         - 예) '데미안, 헤르만 헤세' → 제목=데미안, 저자=헤르만 헤세
         - 예) '진로를 정하지 못한 나, 비정상 인가요?' → 통째로 제목(저자 없음)
      4) 그 외에는 한 줄 전체가 제목
    """
    paren_re = re.compile(r"^(.*\S)\s*[\(（]\s*([^()（）]+?)\s*[\)）]\s*$")

    rows = []
    for line in str(text).splitlines():
        line = line.strip()
        if not line:
            continue

        title, author = line, ""
        if "\t" in line:
            parts = line.split("\t", 1)
            title = parts[0].strip()
            author = parts[1].strip() if len(parts) > 1 else ""
        else:
            m = paren_re.match(line)
            if m and _looks_like_author(m.group(2)):
                # '제목(저자)' 형식
                title, author = m.group(1).strip(), m.group(2).strip()
            elif "," in line:
                head, tail = line.split(",", 1)
                if _looks_like_author(tail):
                    title, author = head.strip(), tail.strip()
                # 저자처럼 안 보이면 line 전체가 제목(기본값 유지)

        if title:
            rows.append({"책 제목": title, "저자": author})
    return pd.DataFrame(rows)


def run_verification(df, title_col, author_col, cert_key, key_suffix):
    """검증 루프 → 통계 → 결과 표 → 엑셀 다운로드. (두 탭에서 공통 사용)"""
    if not cert_key:
        st.error("먼저 좌측 사이드바에서 발급키를 입력해 주세요.")
        return

    total = len(df)
    progress = st.progress(0.0, text="검증을 시작합니다...")
    results = []
    for _, row in df.iterrows():
        title = row[title_col]
        author = "" if (author_col is None or author_col == "(없음)") else row[author_col]
        results.append(verify_row(title, author, cert_key))
        progress.progress(
            len(results) / total,
            text=f"검증 중... ({len(results)}/{total})  현재: {str(title)[:20]}",
        )
        time.sleep(REQUEST_DELAY)
    progress.empty()

    result_df = pd.DataFrame(results, index=df.index)
    merged = pd.concat([df, result_df], axis=1)
    _render_results(merged, result_df, total, key_suffix)


def _render_results(merged, result_df, total, key_suffix):
    """통계 + 결과 표 + 엑셀 다운로드 (여러 탭 공통)."""
    success_cnt = int((result_df["검증결과"] == "성공").sum())
    fail_cnt = int((result_df["검증결과"] == "실패").sum())
    error_cnt = int((result_df["검증결과"] == "오류").sum())

    if error_cnt and error_cnt >= total * 0.5:
        first_err = result_df.loc[result_df["검증결과"] == "오류", "정제된 도서정보"].iloc[0]
        st.error(
            "대부분의 항목이 API 호출 단계에서 막혔습니다. "
            "도서 문제가 아니라 발급키 문제일 가능성이 큽니다.\n\n"
            f"첫 오류 메시지: {first_err}"
        )
        st.caption("좌측 사이드바의 '🔌 연결 테스트'로 키 상태를 먼저 확인해 보세요.")

    st.subheader("📊 검증 통계")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("전체", f"{total} 건")
    m2.metric("성공 ✅", f"{success_cnt} 건")
    m3.metric("실패 ❌", f"{fail_cnt} 건")
    m4.metric("오류 ⚠️", f"{error_cnt} 건")

    st.subheader("📋 검증 결과")
    st.dataframe(merged, use_container_width=True)

    st.download_button(
        label="⬇️ 검증 결과 엑셀 다운로드",
        data=build_excel(merged),
        file_name="독서기록_검증결과.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        key=f"download_{key_suffix}",
    )
    st.success("검증이 완료되었습니다. 위 버튼으로 결과 파일을 내려받으세요.")


def run_isbn_verification(isbns, cert_key, key_suffix):
    """ISBN 목록을 검증."""
    if not cert_key:
        st.error("먼저 좌측 사이드바에서 발급키를 입력해 주세요.")
        return
    total = len(isbns)
    progress = st.progress(0.0, text="검증을 시작합니다...")
    results = []
    for code in isbns:
        results.append(verify_isbn(code, cert_key))
        progress.progress(len(results) / total,
                          text=f"검증 중... ({len(results)}/{total})  ISBN: {code}")
        time.sleep(REQUEST_DELAY)
    progress.empty()

    base = pd.DataFrame({"입력 ISBN": isbns})
    result_df = pd.DataFrame(results)
    merged = pd.concat([base, result_df], axis=1)
    _render_results(merged, result_df, total, key_suffix)


# ---------------------------------------------------------------------------
# Streamlit UI
# ---------------------------------------------------------------------------
def main():
    # --- 제목 줄: 제목 + 제작자 배지 ---
    st.markdown(
        "<div style='display:flex;align-items:center;flex-wrap:wrap;gap:12px;margin-bottom:2px;'>"
        "<span style='font-size:34px;font-weight:800;line-height:1.2;'>📚 독서활동기록 ISBN 검증 대시보드</span>"
        "<span style='display:inline-block;padding:5px 12px;border-radius:999px;"
        "background:linear-gradient(135deg,#2E75B6,#5B9BD5);white-space:nowrap;'>"
        "<span style='color:#FFFFFF;font-weight:800;font-size:13px;letter-spacing:.3px;'>made by 임지환</span>"
        "<span style='color:#E6F0FA;font-weight:500;font-size:12px;'> with Claude</span>"
        "</span></div>",
        unsafe_allow_html=True,
    )
    st.caption(
        "독서기록을 국립중앙도서관 서지정보(ISBN) 데이터베이스와 대조하여 "
        "실제 도서만 정제된 형태로 정리합니다."
    )
    st.link_button(
        "🔗 국립중앙도서관 서지정보 바로가기",
        "https://www.nl.go.kr/seoji/",
    )

    # --- 사이드바: API 키 및 옵션 ---
    with st.sidebar:
        st.header("⚙️ 설정")
        stored_key = get_stored_key()
        if stored_key:
            # 안전 저장소(secrets/환경변수)에 키가 있으면 화면에 노출하지 않음
            cert_key = stored_key
            st.success("🔐 인증키가 안전하게 설정되어 있습니다.")
        else:
            # 저장된 키가 없을 때만 직접 입력받음(비밀번호 형태로 가려짐)
            cert_key = st.text_input(
                "국립중앙도서관 발급키 (key)",
                value="",
                type="password",
                help="서지정보(Seoji) Open API 발급키를 입력하세요. 배포 시 Secrets에 넣으면 이 칸이 사라집니다.",
            ).strip()

        st.markdown("---")
        if st.button("🔌 연결 테스트"):
            if not cert_key:
                st.error("먼저 발급키를 입력하세요.")
            else:
                # 반드시 소장되어 있는 책으로 테스트
                test = call_seoji_api("토지", cert_key)
                if test["error"]:
                    st.error(f"실패: {test['error']}")
                    st.caption("→ 키가 틀렸거나 아직 승인 전일 가능성이 큽니다.")
                elif test["docs"]:
                    st.success(f"성공! 검색 결과 {len(test['docs'])}건을 받았습니다.")
                    sample = test["docs"][0]
                    st.caption(
                        f"예: {strip_html(sample.get('TITLE', '?'))} / "
                        f"{clean_author(sample.get('AUTHOR', '?'))} / "
                        f"ISBN {sample.get('EA_ISBN') or '없음'}"
                    )
                else:
                    st.warning("호출은 됐지만 결과가 0건입니다.")
                # 응답 원문(문제 진단용)
                if test.get("raw"):
                    with st.expander("응답 원문 보기(진단용)"):
                        st.code(test["raw"], language="json")

        st.markdown("---")
        st.markdown(
            "**발급키 안내:** 국립중앙도서관 "
            "[Open API](https://www.nl.go.kr/NL/contents/N31101010000.do) 신청·관리"
        )

    # --- 본문: 세 가지 입력 방식 ---
    tab_paste, tab_excel, tab_isbn = st.tabs(["✍️ 직접 입력", "📂 엑셀 업로드", "🔢 ISBN"])

    # ===== 탭 1: 엑셀 업로드 =====
    with tab_excel:
        uploaded = st.file_uploader(
            "독서기록 엑셀 파일 업로드 (.xlsx)", type=["xlsx"], accept_multiple_files=False
        )
        if uploaded is None:
            st.info("좌측 설정에서 발급키를 확인하고, 위에 엑셀 파일을 업로드해 주세요.")
        else:
            try:
                df = pd.read_excel(uploaded)
            except Exception as e:  # noqa: BLE001
                st.error(f"엑셀을 읽는 중 오류가 발생했습니다: {e}")
                df = None

            if df is not None and df.empty:
                st.warning("업로드한 파일에 데이터가 없습니다.")
            elif df is not None:
                st.subheader("📄 업로드 데이터 미리보기")
                st.dataframe(df.head(10), use_container_width=True)

                cols = list(df.columns)

                def guess(candidates):
                    for i, c in enumerate(cols):
                        if any(k in str(c) for k in candidates):
                            return i
                    return 0

                c1, c2 = st.columns(2)
                with c1:
                    title_col = st.selectbox(
                        "‘책 제목’ 열 선택", cols,
                        index=guess(["제목", "책", "도서", "title"]),
                    )
                with c2:
                    author_options = ["(없음)"] + cols
                    a_guess = guess(["저자", "작가", "author"])
                    author_col = st.selectbox(
                        "‘저자’ 열 선택 (선택)", author_options,
                        index=(a_guess + 1) if a_guess else 0,
                    )

                if st.button("🔍 ISBN 검증 시작", type="primary", key="verify_excel"):
                    run_verification(df, title_col, author_col, cert_key, "excel")

    # ===== 탭 2: 직접 입력 =====
    with tab_paste:
        st.caption(
            "한 줄에 한 권씩 입력하세요. 저자를 함께 적으면 같은 제목의 다른 책과 더 정확히 구분됩니다. "
            "‘제목(저자)’ 또는 ‘제목, 저자’ 형식 모두 가능하며 저자는 생략할 수 있습니다. "
            "제목에 쉼표·물음표가 들어간 책은 통째로 제목으로 처리되니, 그런 책은 엑셀에서 두 열을 복사해 "
            "붙여넣으면(탭 구분) 가장 정확합니다."
        )
        sample = "커튼콜(조우리)\n데미안, 헤르만 헤세\n어린 왕자\n소년이 온다(한강)"
        text = st.text_area(
            "도서 목록 붙여넣기",
            height=200,
            placeholder=sample,
            key="paste_area",
        )
        if st.button("🔍 ISBN 검증 시작", type="primary", key="verify_paste"):
            df_paste = parse_pasted(text)
            if df_paste.empty:
                st.warning("입력된 도서가 없습니다. 한 줄에 한 권씩 입력해 주세요.")
            else:
                st.caption(f"{len(df_paste)}권을 검증합니다.")
                run_verification(df_paste, "책 제목", "저자", cert_key, "paste")

    # ===== 탭 3: ISBN =====
    with tab_isbn:
        st.caption(
            "ISBN을 한 줄에 하나씩 입력하세요. 하이픈(-)이 있어도 됩니다. "
            "ISBN으로 찾으면 권수·판본까지 정확한 책 한 권을 바로 확인합니다."
        )
        isbn_text = st.text_area(
            "ISBN 목록 붙여넣기",
            height=200,
            placeholder="9791136401212\n978-89-364-7423-5\n9788937460777",
            key="isbn_area",
        )
        if st.button("🔍 ISBN 검증 시작", type="primary", key="verify_isbn"):
            codes = []
            for line in str(isbn_text).splitlines():
                c = re.sub(r"[^0-9Xx]", "", line)
                if c:
                    codes.append(c)
            if not codes:
                st.warning("입력된 ISBN이 없습니다. 한 줄에 하나씩 입력해 주세요.")
            else:
                st.caption(f"{len(codes)}건을 검증합니다.")
                run_isbn_verification(codes, cert_key, "isbn")


if __name__ == "__main__":
    main()
